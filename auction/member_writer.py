"""Write extra members back to the member workbook."""

from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
import sys

import openpyxl

from .identity import normalize_id


@dataclass(frozen=True)
class MemberWriteResult:
    action: str
    row: int


def _find_header_row(sheet, required_headers: set[str], max_scan_rows: int = 20) -> tuple[int, list[str]]:
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        headers = [str(header).strip() if header else "" for header in row]
        if required_headers.issubset(set(headers)):
            return row_number, headers
    return 1, []


def _ensure_workbook(path: Path, headers: list[str]):
    if path.exists():
        return openpyxl.load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "회원명단"
    sheet.append(headers)
    return workbook


def upsert_member(
    member_file: str,
    *,
    display_id: str,
    customer_name: str,
    phone: str,
    address: str,
    col_id: str = "아이디",
    col_name: str = "고객명",
    col_phone: str = "연락처",
    col_address: str = "주소",
) -> MemberWriteResult:
    """Append or update one member in the selected Excel member file."""
    path = Path(member_file).expanduser()
    if not path:
        raise ValueError("회원명단 엑셀 파일을 먼저 선택하세요.")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("회원명단은 .xlsx 또는 .xlsm 파일이어야 합니다.")

    required_headers = [col_id, col_name, col_phone, col_address]
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = _ensure_workbook(path, required_headers)
    sheet = workbook.active

    header_row, headers = _find_header_row(sheet, set(required_headers))
    if not headers:
        header_row = 1
        for column, header in enumerate(required_headers, start=1):
            sheet.cell(row=header_row, column=column).value = header
        headers = required_headers[:]

    def column_index(name: str) -> int:
        if name not in headers:
            headers.append(name)
            sheet.cell(row=header_row, column=len(headers)).value = name
        return headers.index(name) + 1

    id_col = column_index(col_id)
    name_col = column_index(col_name)
    phone_col = column_index(col_phone)
    address_col = column_index(col_address)

    key = normalize_id(display_id)
    target_row = None
    for row_number in range(header_row + 1, sheet.max_row + 1):
        raw_id = sheet.cell(row=row_number, column=id_col).value
        if raw_id and normalize_id(str(raw_id)) == key:
            target_row = row_number
            break

    action = "updated" if target_row else "created"
    if target_row is None:
        target_row = sheet.max_row + 1

    sheet.cell(row=target_row, column=id_col).value = display_id
    sheet.cell(row=target_row, column=name_col).value = customer_name
    sheet.cell(row=target_row, column=phone_col).value = phone
    sheet.cell(row=target_row, column=address_col).value = address
    workbook.save(path)
    workbook.close()

    return MemberWriteResult(action=action, row=target_row)


def main() -> None:
    if len(sys.argv) != 2:
        print("사용법: python3 -m auction.member_writer payload.json", file=sys.stderr)
        sys.exit(2)

    payload = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    result = upsert_member(
        payload.get("member_file", ""),
        display_id=str(payload.get("display_id", "")).strip(),
        customer_name=str(payload.get("customer_name", "")).strip(),
        phone=str(payload.get("phone", "")).strip(),
        address=str(payload.get("address", "")).strip(),
        col_id=payload.get("col_id", "아이디"),
        col_name=payload.get("col_name", "고객명"),
        col_phone=payload.get("col_phone", "연락처"),
        col_address=payload.get("col_address", "주소"),
    )
    print(json.dumps({"action": result.action, "row": result.row}, ensure_ascii=False), flush=True)


if __name__ == "__main__":
    main()
