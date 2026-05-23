"""Excel order workbook creation and row writing."""

from __future__ import annotations

from dataclasses import dataclass

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .members import Member


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


FILLS = {
    "header": _fill("2F4F8F"),
    "no_member": _fill("FFCCCC"),
    "blocked": _fill("FF4444"),
    "duplicate": _fill("FFFF99"),
    "auction": _fill("E8F4FD"),
    "first": _fill("E8FDE8"),
}

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER_COLS = {1, 2, 4, 5, 6, 11, 12, 13, 14}

HEADERS = [
    "시간", "번호", "주문내역", "금액", "수량", "총금액",
    "아이디", "고객명", "연락처", "주소",
    "입금확인", "방문수령", "합배송", "반품", "비고",
]
COL_WIDTHS = [10, 5, 42, 9, 5, 9, 22, 11, 15, 36, 8, 8, 8, 6, 22]


@dataclass
class OrderRow:
    time_text: str
    sequence: int
    product: str
    price: int
    quantity: int
    uid: str
    member: Member
    is_member: bool
    is_blocked: bool
    is_duplicate: bool
    notes: str
    auction_type: str

    @property
    def total(self) -> int:
        return (self.price or 0) * (self.quantity or 0)

    def values(self) -> list:
        fields = self.member.to_order_fields()
        return [
            self.time_text,
            self.sequence,
            self.product,
            self.price,
            self.quantity,
            self.total,
            self.uid,
            fields["고객명"],
            fields["연락처"],
            fields["주소"],
            "",
            "",
            "",
            "",
            self.notes,
        ]

    def fill_key(self) -> str:
        if self.is_blocked:
            return "blocked"
        if not self.is_member:
            return "no_member"
        if self.is_duplicate:
            return "duplicate"
        if self.auction_type == "first_come":
            return "first"
        return "auction"


def init_workbook():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "주문서"
    header_font = Font(color="FFFFFF", bold=True, size=10, name="맑은 고딕")
    for index, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.fill = FILLS["header"]
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 20
    sheet.freeze_panes = "A2"
    return workbook, sheet


def write_order_row(sheet, row_num: int, order: OrderRow) -> None:
    row_fill = FILLS[order.fill_key()]
    body_font = Font(name="맑은 고딕", size=9)
    for column, value in enumerate(order.values(), 1):
        cell = sheet.cell(row=row_num, column=column, value=value)
        cell.fill = row_fill
        cell.font = body_font
        cell.border = BORDER
        cell.alignment = Alignment(
            vertical="center",
            horizontal="center" if column in CENTER_COLS else "left",
        )
    sheet.row_dimensions[row_num].height = 16


def sort_and_renumber(sheet, orders: list[OrderRow]) -> None:
    """Sort order rows by Korean customer name and rewrite row numbers."""
    if not orders:
        return

    orders.sort(key=lambda order: order.member.customer_name or "힣힣힣")
    for index, order in enumerate(orders, start=1):
        order.sequence = index
        write_order_row(sheet, index + 1, order)
