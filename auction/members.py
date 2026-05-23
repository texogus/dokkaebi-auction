"""Excel loaders for members and blocked users."""

from __future__ import annotations

from dataclasses import dataclass

import openpyxl

from .config import Settings
from .identity import normalize_id


@dataclass(frozen=True)
class Member:
    customer_name: str = ""
    phone: str = ""
    address: str = ""

    def to_order_fields(self) -> dict[str, str]:
        return {
            "고객명": self.customer_name,
            "연락처": self.phone,
            "주소": self.address,
        }


def _column_index(headers: list[str], name: str) -> int:
    return headers.index(name) if name in headers else -1


def _find_header_row(sheet, required_headers: set[str], max_scan_rows: int = 20) -> tuple[int, list[str]]:
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        headers = [str(header).strip() if header else "" for header in row]
        if required_headers.issubset(set(headers)):
            return row_number, headers
    return 1, []


def load_members(settings: Settings) -> dict[str, Member]:
    """Load members keyed by normalized YouTube display name."""
    members: dict[str, Member] = {}
    try:
        workbook = openpyxl.load_workbook(settings.member_file, read_only=True, data_only=True)
        sheet = workbook.active
        required_headers = {settings.col_id, settings.col_name, settings.col_phone, settings.col_address}
        header_row, headers = _find_header_row(sheet, required_headers)
        if not headers:
            raise ValueError(f"필수 헤더를 찾지 못했습니다: {', '.join(sorted(required_headers))}")

        id_col = _column_index(headers, settings.col_id)
        name_col = _column_index(headers, settings.col_name)
        phone_col = _column_index(headers, settings.col_phone)
        address_col = _column_index(headers, settings.col_address)

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            raw_id = row[id_col] if id_col >= 0 else None
            if not raw_id:
                continue
            members[normalize_id(str(raw_id))] = Member(
                customer_name=str(row[name_col]).strip() if name_col >= 0 and row[name_col] else "",
                phone=str(row[phone_col]).strip() if phone_col >= 0 and row[phone_col] else "",
                address=str(row[address_col]).strip() if address_col >= 0 and row[address_col] else "",
            )
        workbook.close()
        print(f"회원 {len(members):,}명 로드 완료")
    except FileNotFoundError:
        print(f"회원 명단({settings.member_file}) 없음: 대조 생략")
    except Exception as exc:
        print(f"회원 명단 오류: {exc}")
    return members


def load_blocked(settings: Settings) -> set[str]:
    """Load blocked display names keyed by normalized id."""
    blocked: set[str] = set()
    try:
        workbook = openpyxl.load_workbook(settings.blocked_file, read_only=True, data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                blocked.add(normalize_id(str(row[0])))
        workbook.close()
        print(f"차단자 {len(blocked)}명 로드 완료")
    except FileNotFoundError:
        pass
    except Exception as exc:
        print(f"차단 목록 오류: {exc}")
    return blocked
