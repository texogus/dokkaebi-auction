import tempfile
from pathlib import Path
import unittest

import openpyxl

from auction.member_writer import upsert_member


class MemberWriterTest(unittest.TestCase):
    def test_upsert_member_appends_to_member_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "members.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["아이디", "고객명", "연락처", "주소"])
            workbook.save(path)
            workbook.close()

            result = upsert_member(
                str(path),
                display_id="prexogus",
                customer_name="기대현",
                phone="01086828102",
                address="일산서구 위브더제니스 107동 501호",
            )

            self.assertEqual(result.action, "created")
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            self.assertEqual(sheet.cell(row=2, column=1).value, "prexogus")
            self.assertEqual(sheet.cell(row=2, column=2).value, "기대현")
            workbook.close()

    def test_upsert_member_updates_existing_member_without_duplicate(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "members.xlsx"
            upsert_member(str(path), display_id="@prexogus", customer_name="기대현", phone="", address="")

            result = upsert_member(
                str(path),
                display_id="prexogus",
                customer_name="기대현2",
                phone="010",
                address="주소",
            )

            self.assertEqual(result.action, "updated")
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(sheet.cell(row=2, column=1).value, "prexogus")
            self.assertEqual(sheet.cell(row=2, column=2).value, "기대현2")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
